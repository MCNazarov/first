from ldap3 import Server, Connection, SIMPLE, SYNC, ASYNC, SUBTREE, ALL
import openpyxl
# домен - example.com
# DNS имя сервера Active Directory

# Пользователь (логин) в Active Directory - нужно указать логин в AD
# в формате 'EXAMPLE\aduser' или 'aduser@example.com'
from openpyxl.worksheet.worksheet import Worksheet
from Work_Exel import open_fail

AD_SEARCH_TREE = 'dc=xxxx,dc=XXXX'
print(type(AD_SEARCH_TREE))
# указваем имя домена
print('Введити имя домена')
AD_SERVER = input()
print(type(AD_SERVER))
# логин пользователя с правами доступа к домену
print('Введити логин пользователя')
AD_USER = input() + '@' + AD_SERVER + '.ru'
# пароль пользователя домена
print('Введити пароль пользователя', AD_USER)
AD_PASSWORD = input()
# задаем дерево для поиска
AD_SEARCH_TREE = 'dc=' + AD_SERVER + ',dc=ru'
AD_SERVER = AD_SERVER + '.ru'

# проверяем введенные данные
print(AD_SERVER, AD_USER, AD_PASSWORD,AD_SEARCH_TREE)
#AD_PASSWORD = input()

# создаем файл в который будет записывться результат работы скрипта
open_fail()


server = Server(AD_SERVER)
conn = Connection(server,user=AD_USER,password=AD_PASSWORD)
conn.bind()
# в ответ должно быть - True

# Поиск в Active Directory
# примеры ldap фильтров можно посмотреть здесь -
# https://social.technet.microsoft.com/wiki/contents/articles/8077.active-directory-ldap-ru-ru.aspx
# Я в нижеследующем фильтре:
# - исключаю всеx отключенных пользователей (!(UserAccountControl:1.2.840.113556.1.4.803:=2))
# - добавляю только тех пользователей у которых заполнено имя и фамилия
# - и вывожу атрибуты - attributes
# Все возможные атрибуты Active Directory можно посмотреть здесь -
# https://msdn.microsoft.com/en-us/library/ms675090%28v=vs.85%29.aspx
conn.search(AD_SEARCH_TREE,'(&(objectCategory=Person)(!(UserAccountControl:1.2.840.113556.1.4.803:=2))(givenName=*)(sn=*))',
    SUBTREE,
    attributes =['cn','proxyAddresses','department','sAMAccountName', 'displayName', 'telephoneNumber', 'ipPhone', 'streetAddress',
    'title','manager','objectGUID','company','lastLogon']
    )
# после этого запроса в ответ должно быть - True
# или вывести только Common-Name - cn
# функция записи пользователей из АД в вайл xlsx
def NameSheetFile(namelist):
    wb = openpyxl.load_workbook('example.xlsx')
    title = str(namelist)
    wb.create_sheet(title=title, index=0)
    wb.save('example.xlsx')

def  writeInFile(locentry,user):
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb.active
    index = 'A' + str(locentry)
    sheet[index] = str(user)
    wb.save('example.xlsx')

n=0 # счетчик пользователей в АД
NameSheetFile('Users')
for entry in conn.entries:
    n=n+1
    print(n, entry.cn)
    writeInFile(n, entry.cn)


conn.search(AD_SEARCH_TREE,'(objectCategory=group)')
NameSheetFile('Grupp')
p=0 # счетчик пользователей в АД
for entry in conn.entries:
    p=p+1
    print(p, entry)
    writeInFile(p, entry)



# Найти пользователя с логином admin (sAMAccountName=admin) и показать информацию по нему
#conn.search(AD_SEARCH_TREE,'(&(objectCategory=Person)(sAMAccountName=nazarov))', SUBTREE,
#    attributes =['cn','proxyAddresses','department','sAMAccountName', 'displayName', 'telephoneNumber', 'ipPhone', 'streetAddress',
#    'title','manager','objectGUID','company','lastLogon']
#    )

conn.entries